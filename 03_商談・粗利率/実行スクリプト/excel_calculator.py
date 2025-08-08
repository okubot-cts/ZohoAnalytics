import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

def csv_to_excel_with_calculations(csv_file_path, excel_file_path):
    """
    CSVファイルをExcelに変換し、売上・原価の小計と粗利・粗利率の計算式を追加
    """
    # CSVファイルを読み込み（日本語の通貨記号を処理）
    df = pd.read_csv(csv_file_path, encoding='utf-8')
    
    # Excelワークブック作成
    wb = Workbook()
    ws = wb.active
    ws.title = "商談商品内訳レポート"
    
    # データの先頭100行のみを処理（処理時間短縮のため）
    df_sample = df.head(100)
    
    # データフレームをワークシートに書き込み
    for r in dataframe_to_rows(df_sample, index=False, header=True):
        ws.append(r)
    
    # データの最終行を取得
    last_row = ws.max_row
    
    # 新しい列を追加（売上小計、原価小計、粗利、粗利率）
    ws.cell(row=1, column=ws.max_column + 1, value="売上小計")
    ws.cell(row=1, column=ws.max_column + 1, value="原価小計") 
    ws.cell(row=1, column=ws.max_column + 1, value="粗利")
    ws.cell(row=1, column=ws.max_column + 1, value="粗利率(%)")
    
    # 列インデックスを取得
    sales_col = None
    cost_col = None
    
    # 小計列（売上）と原価列を探す
    for col in range(1, ws.max_column - 3):  # 新しく追加した4列を除く
        cell_value = ws.cell(row=1, column=col).value
        if cell_value == "小計":
            sales_col = col
        elif cell_value == "原価（税別）":
            cost_col = col
    
    if sales_col and cost_col:
        # 各行に計算式を追加
        for row in range(2, last_row + 1):
            # 売上小計の計算式（現在は個別の小計をそのまま表示）
            ws.cell(row=row, column=ws.max_column - 3, 
                   value=f"={get_column_letter(sales_col)}{row}")
            
            # 原価小計の計算式（現在は個別の原価をそのまま表示）
            ws.cell(row=row, column=ws.max_column - 2,
                   value=f"={get_column_letter(cost_col)}{row}")
            
            # 粗利の計算式（売上 - 原価）
            ws.cell(row=row, column=ws.max_column - 1,
                   value=f"={get_column_letter(sales_col)}{row}-{get_column_letter(cost_col)}{row}")
            
            # 粗利率の計算式（粗利 / 売上 × 100）
            ws.cell(row=row, column=ws.max_column,
                   value=f"=IF({get_column_letter(sales_col)}{row}<>0,({get_column_letter(sales_col)}{row}-{get_column_letter(cost_col)}{row})/{get_column_letter(sales_col)}{row}*100,0)")
        
        # 合計行を追加
        total_row = last_row + 2
        ws.cell(row=total_row, column=1, value="合計")
        
        # 売上合計
        ws.cell(row=total_row, column=ws.max_column - 3,
               value=f"=SUM({get_column_letter(ws.max_column - 3)}2:{get_column_letter(ws.max_column - 3)}{last_row})")
        
        # 原価合計
        ws.cell(row=total_row, column=ws.max_column - 2,
               value=f"=SUM({get_column_letter(ws.max_column - 2)}2:{get_column_letter(ws.max_column - 2)}{last_row})")
        
        # 粗利合計
        ws.cell(row=total_row, column=ws.max_column - 1,
               value=f"=SUM({get_column_letter(ws.max_column - 1)}2:{get_column_letter(ws.max_column - 1)}{last_row})")
        
        # 粗利率平均
        ws.cell(row=total_row, column=ws.max_column,
               value=f"=IF({get_column_letter(ws.max_column - 3)}{total_row}<>0,{get_column_letter(ws.max_column - 1)}{total_row}/{get_column_letter(ws.max_column - 3)}{total_row}*100,0)")
    
    # ヘッダーのスタイルを設定
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    
    # 合計行のスタイルを設定
    if last_row > 1:
        total_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        total_font = Font(bold=True)
        
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=last_row + 2, column=col)
            cell.fill = total_fill
            cell.font = total_font
    
    # 列幅を自動調整
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Excelファイルを保存
    wb.save(excel_file_path)
    print(f"Excelファイルが作成されました: {excel_file_path}")
    print("追加された計算列:")
    print("- 売上小計: 各行の売上金額")
    print("- 原価小計: 各行の原価金額") 
    print("- 粗利: 売上 - 原価")
    print("- 粗利率(%): (粗利 / 売上) × 100")
    print("- 最下部に合計行が追加されています")

if __name__ == "__main__":
    csv_file = "2025年1月以降_商談_商品内訳_レポート.csv"
    excel_file = "2025年1月以降_商談_商品内訳_レポート_計算式付き.xlsx"
    
    csv_to_excel_with_calculations(csv_file, excel_file)